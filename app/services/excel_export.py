from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models.receipt import StoredReceipt

CURRENCY_FORMAT = "#,##0.00"
HEADERS = [
    "product_name",
    "price",
    "quantity",
    "price_per_item",
    "user_id",
    "date",
    "shop_name",
]


def compute_receipt_rows(receipts: list[StoredReceipt]) -> tuple[list[tuple], float]:
    """Pure function, no openpyxl involved. Returns (rows, grand_total)."""
    rows = []
    for receipt in receipts:
        for item in receipt.items:
            rows.append(
                (
                    item.name,  # product_name
                    item.total_price,  # price
                    item.quantity,  # quantity
                    item.unit_price,  # price_per_item
                    receipt.user_id,  # user_id
                    receipt.date,  # date
                    receipt.store_name,  # shop_name
                )
            )
    grand_total = sum(receipt.total for receipt in receipts)
    return rows, grand_total


def _write_workbook(receipts: list[StoredReceipt]) -> bytes:
    rows, grand_total = compute_receipt_rows(receipts)

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Receipts"

    sheet.append(HEADERS)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(list(row))
        sheet.cell(row=sheet.max_row, column=2).number_format = CURRENCY_FORMAT
        sheet.cell(row=sheet.max_row, column=4).number_format = CURRENCY_FORMAT

    # sheet.append([])
    # sheet.append(["Grand Total", grand_total])
    # sheet.cell(row=sheet.max_row, column=2).number_format = CURRENCY_FORMAT
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    for index, column_cells in enumerate(sheet.columns, start=1):
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value not in (None, "")]
        width = max(lengths, default=8) + 2
        sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def receipt_to_excel(receipt: StoredReceipt) -> bytes:
    return _write_workbook([receipt])


def combine_receipts_to_excel(receipts: list[tuple[str, StoredReceipt]]) -> bytes:
    return _write_workbook([r for _id, r in receipts])
