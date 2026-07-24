from io import BytesIO

import openpyxl

from app.models.receipt import ReceiptItem, StoredReceipt
from app.services.excel_export import (
    HEADERS,
    combine_receipts_to_excel,
    compute_receipt_rows,
    receipt_to_excel,
)


def _make_receipt(
    store_name: str = "Aldi",
    date: str = "2026-07-20",
    user_id: str = "default",
) -> StoredReceipt:
    items = [
        ReceiptItem(name="Milk", quantity=2, unit_price=1.5, total_price=3.0),
        ReceiptItem(name="Bread", quantity=1, unit_price=2.5, total_price=2.5),
    ]
    return StoredReceipt(
        store_name=store_name,
        date=date,
        items=items,
        subtotal=5.5,
        tax_amount=0.5,
        total=6.0,
        user_id=user_id,
    )


def test_receipt_to_excel_matches_fixture():
    receipt = _make_receipt()

    excel_bytes = receipt_to_excel(receipt)
    workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
    sheet = workbook.active
    assert sheet is not None

    header_row = [cell.value for cell in sheet[1]]
    assert header_row == HEADERS

    expected_rows = [
        (
            item.name,
            item.total_price,
            item.quantity,
            item.unit_price,
            receipt.user_id,
            receipt.date,
            receipt.store_name,
        )
        for item in receipt.items
    ]
    for row_index, expected in enumerate(expected_rows, start=2):
        actual = [cell.value for cell in sheet[row_index]]
        assert tuple(actual) == expected

    # no grand total row — the sheet ends right after the last item row
    assert sheet.max_row == 1 + len(expected_rows)


def test_compute_receipt_rows_single_receipt():
    receipt = _make_receipt()

    rows, grand_total = compute_receipt_rows([receipt])

    assert rows == [
        ("Milk", 3.0, 2, 1.5, "default", "2026-07-20", "Aldi"),
        ("Bread", 2.5, 1, 2.5, "default", "2026-07-20", "Aldi"),
    ]
    assert grand_total == receipt.total


def test_compute_receipt_rows_multiple_receipts():
    receipt_a = _make_receipt(store_name="Aldi", date="2026-07-20", user_id="alice")
    receipt_b = _make_receipt(store_name="Rewe", date="2026-07-21", user_id="bob")

    rows, grand_total = compute_receipt_rows([receipt_a, receipt_b])

    assert len(rows) == 4
    assert rows[:2] == [
        ("Milk", 3.0, 2, 1.5, "alice", "2026-07-20", "Aldi"),
        ("Bread", 2.5, 1, 2.5, "alice", "2026-07-20", "Aldi"),
    ]
    assert rows[2:] == [
        ("Milk", 3.0, 2, 1.5, "bob", "2026-07-21", "Rewe"),
        ("Bread", 2.5, 1, 2.5, "bob", "2026-07-21", "Rewe"),
    ]
    assert grand_total == receipt_a.total + receipt_b.total


def test_combine_receipts_to_excel_only_includes_selected():
    receipt_a = _make_receipt(store_name="Aldi", date="2026-07-20", user_id="alice")
    receipt_b = _make_receipt(store_name="Rewe", date="2026-07-21", user_id="bob")
    receipt_c = _make_receipt(store_name="Lidl", date="2026-07-22", user_id="carol")

    excel_bytes = combine_receipts_to_excel(
        [("id-a", receipt_a), ("id-b", receipt_b)]
    )

    workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
    sheet = workbook.active
    assert sheet is not None

    shop_names = {str(row[6].value) for row in sheet.iter_rows(min_row=2, max_row=5)}
    assert shop_names == {"Aldi", "Rewe"}
    assert receipt_c.store_name not in shop_names

    # no grand total row — 1 header + 4 item rows (2 receipts x 2 items), nothing after
    assert sheet.max_row == 5
